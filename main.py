from twilio.rest import Client
from openpyxl import load_workbook
import configparser
config = configparser.ConfigParser()

config.read('config.ini')

# Twilio account credentials
account_sid = config.get('Twilio', 'account_sid')
auth_token = config.get('Twilio', 'auth_token')
twilio_number = config.get('Twilio', 'twilio_number')

# Load Excel sheet
excel_file = 'contacts.xlsx'
wb = load_workbook(excel_file)
sheet = wb.active

# Initialize Twilio client
client = Client(account_sid, auth_token)

# Iterate over each row in the Excel sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    phone_number = row[0]  # Assuming phone numbers are in the first column
    message = row[1]      # Assuming messages are in the second column

    # Send WhatsApp message
    message = client.messages.create(
        from_='whatsapp:' + twilio_number,
        body=message,
        to='whatsapp:' + phone_number
    )

    print(f'Message sent to {phone_number}: {message.sid}')
